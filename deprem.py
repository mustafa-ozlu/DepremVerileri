import pandas as pd
import requests
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import os

# Sarı dolgu tanımı
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF2400", end_color="FF2400", fill_type="solid")

# Web sayfasından veriyi çekme
url = "http://www.koeri.boun.edu.tr/scripts/lst6.asp"
response = requests.get(url)
response.encoding = "ISO-8859-9"  # Türkçe karakter desteği için

data = response.text

# İlgili kısmı ayıklama
start_index = data.find("..................TÜRKİYE VE YAKIN ÇEVRESİNDEKİ SON DEPREMLER....................")
if start_index != -1:
    data = data[start_index:]
    lines = data.split("\n")[6:]  # Başlıkları atlamak için ilk birkaç satırı atlıyoruz
else:
    lines = []

parsed_data = []
for line in lines[:250]:
    parts = line.split()
    if len(parts) < 10:
        continue
    tarih, saat = parts[0], parts[1]
    enlem, boylam, derinlik = parts[2], parts[3], parts[4]
    md, ml, mw = parts[5], parts[6], parts[7]
    yer = " ".join(parts[8:-1])  # Yer adını al
    #cozum_niteligi = parts[-1]    # Çözüm niteliği
    harita=f'=HYPERLINK("https://www.openstreetmap.org/?mlat={enlem}&mlon={boylam}&zoom=9","Haritada Gör")'

   
    
    parsed_data.append([tarih, saat, enlem, boylam, derinlik,  ml, mw, yer, harita])

# Verileri Excel'e kaydet
columns = ["Tarih", "Saat", "Enlem", "Boylam", "Derinlik(km)",  "ML", "Mw", "Yer", "Harita"]
df = pd.DataFrame(parsed_data, columns=columns)
excel_path = "deprem_verileri.xlsx"
df.to_excel(excel_path, index=False, engine='openpyxl')

# Excel dosyasını yükle ve sütun genişliklerini ayarla
wb = load_workbook(excel_path)
ws = wb.active

# Sütun genişliklerini otomatik olarak belirle
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Sütun harfini al
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))  # Maksimum içerik uzunluğunu bul
    ws.column_dimensions[col_letter].width = max_length + 2  # Biraz boşluk ekleyerek genişliği ayarla

# Belirli sütunlardaki hücreleri ortala
columns_to_center = ["A", "B", "C", "D", "E", "F", "G", "H","I"]  # Ortalanacak sütunlar
for col in columns_to_center:
    for cell in ws[col]:  # Sütundaki tüm hücreleri ortala
        cell.alignment = Alignment(horizontal="center")
# J sütununun genişliğini "Harita" sütununun genişliği ile eşitle
ws.column_dimensions["I"].width = 20
for row in ws.iter_rows(min_row=2):  # Başlık satırını atla
    ml_cell = row[5]  # G sütunu, 0-indeksli olduğu için 6. hücre
    try:
        value = float(ml_cell.value)
        if 3 <= value < 5:
            for cell in row:
                cell.fill = yellow_fill
        elif value >= 5:
            for cell in row:
                cell.fill = red_fill
    except (TypeError, ValueError):
        pass  # Boş ya da sayıya çevrilemeyen hücreler için
# Güncellenmiş dosyayı kaydet
wb.save(excel_path)

# Excel dosyasını otomatik olarak aç
print("Excel Aciliyor")
os.startfile(excel_path)

